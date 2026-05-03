import streamlit as st
import pdfplumber
import pandas as pd
import io
import pytesseract
import re
import fitz  # PyMuPDF: Reemplaza a pdf2image para evitar el error de Poppler
from PIL import Image
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURACIÓN Y ESTILO UI PREMIUM (GLASSMORPHISM)
# ============================================================
st.set_page_config(page_title="INVIMA Data Engine", layout="wide", page_icon="🏛️")

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    
    .stApp { background: radial-gradient(circle at top left, #fdfbfb, #ebedee); }

    .file-card {
        background: white;
        padding: 2rem;
        border-radius: 12px;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
        border: 1px solid #e2e8f0;
        margin-bottom: 1.5rem;
    }

    .main-title {
        color: #1a365d;
        font-weight: 800;
        text-align: center;
        font-size: 2.5rem;
        margin-bottom: 1rem;
    }

    .stButton>button {
        width: 100%;
        border-radius: 8px;
        background-color: #2b6cb0;
        color: white;
        font-weight: 600;
        border: none;
        padding: 0.6rem;
    }
    .stButton>button:hover { background-color: #2c5282; border: none; color: white; }

    [data-testid="stSidebar"] { background-color: #ffffff; border-right: 1px solid #e2e8f0; }
    </style>
    """, unsafe_allow_html=True)

# ============================================================
# MOTOR DE EXTRACCIÓN (HÍBRIDO + OCR REAL)
# ============================================================

def extraer_con_ocr(file):
    """Convierte PDF a imagen usando PyMuPDF y usa Tesseract para tabular."""
    filas = []
    file.seek(0)
    pdf_bytes = file.read()
    
    # Abrimos el PDF desde los bytes con PyMuPDF (fitz)
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    
    for page in doc:
        # Renderizamos la página a una imagen (pixmap)
        # 300 DPI equivale a una matriz de zoom de 4.166 (300/72)
        zoom = 300 / 72
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        
        # Convertimos el pixmap a una imagen PIL para Tesseract
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        
        # Usamos Tesseract para buscar tablas (--psm 6)
        text = pytesseract.image_to_string(img, lang='spa', config='--psm 6')
        
        for line in text.split('\n'):
            # Detectamos separaciones grandes de espacios para simular columnas
            partes = [p.strip() for p in re.split(r'\t| {2,}', line) if p.strip()]
            if len(partes) > 1:
                filas.append(partes)
    
    doc.close()
    return filas

def extraer_con_limpieza(file, modo_ocr=False):
    if modo_ocr:
        return extraer_con_ocr(file)
    
    filas = []
    file.seek(0)
    try:
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                # Intentamos con estructura de líneas
                table = page.extract_table({"vertical_strategy": "lines", "horizontal_strategy": "lines"})
                if not table:
                    table = page.extract_table()
                
                if table:
                    for row in table:
                        clean_row = [" ".join(str(c).split()) if c else "" for c in row]
                        if any(clean_row): filas.append(clean_row)
    except:
        return extraer_con_ocr(file) # Fallback automático
    
    # Si pdfplumber no encontró nada, intentamos OCR
    if not filas:
        return extraer_con_ocr(file)
        
    return filas

# ============================================================
# LÓGICA DE EXPORTACIÓN (FORMATO PROFESIONAL)
# ============================================================

def generar_excel_profesional(final_data):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for name, df in final_data.items():
            sheet_name = name[:30]
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            ws = writer.sheets[sheet_name]
            azul_invima = PatternFill(start_color="002F6C", end_color="002F6C", fill_type="solid")
            fuente_blanca = Font(color="FFFFFF", bold=True, size=11)
            
            ws.sheet_view.showGridLines = False
            
            for col_num, value in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = azul_invima
                cell.font = fuente_blanca
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-ajuste de columnas con seguridad
                try:
                    max_val = df[value].astype(str).map(len).max()
                    max_length = max(max_val, len(str(value))) + 2
                except:
                    max_length = 20
                ws.column_dimensions[get_column_letter(col_num)].width = min(max_length, 50)

    return output.getvalue()

# ============================================================
# INTERFAZ DE USUARIO
# ============================================================

st.markdown("<h1 class='main-title'>🏛️ INVIMA DATA ENGINE</h1>", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### 🛠️ Configuración")
    modo_scan = st.toggle("🔦 Activar Motor OCR", help="Activa esto si el PDF es un escaneo de baja calidad.")
    if st.button("🔄 Limpiar Sesión"):
        st.session_state.clear()
        st.rerun()

uploaded_files = st.file_uploader("Cargar Reportes PDF", type=["pdf"], accept_multiple_files=True)

if uploaded_files:
    if 'master' not in st.session_state: st.session_state.master = {}

    for i, f in enumerate(uploaded_files):
        with st.container():
            st.markdown('<div class="file-card">', unsafe_allow_html=True)
            c1, c2 = st.columns([4, 1])
            c1.markdown(f"**Archivo:** `{f.name}`")
            
            if c2.button("Analizar", key=f"btn_{i}"):
                with st.spinner("Extrayendo información..."):
                    datos = extraer_con_limpieza(f, modo_scan)
                    if datos:
                        st.session_state.master[f.name] = {"df": pd.DataFrame(datos), "clean": None}
                    else:
                        st.error("No se pudo leer el contenido.")

            if f.name in st.session_state.master:
                res = st.session_state.master[f.name]
                df_raw = res["df"]
                
                h_idx = st.number_input("Selecciona fila de encabezados:", 0, max(0, len(df_raw)-1), 0, key=f"h_{i}")
                st.dataframe(df_raw.head(8), use_container_width=True)
                
                if st.button("💎 Generar Estructura Limpia", key=f"fix_{i}"):
                    df_final = df_raw.iloc[h_idx:].copy()
                    df_final.columns = [str(c).strip() for c in df_final.iloc[0]]
                    df_final = df_final[1:].reset_index(drop=True)
                    st.session_state.master[f.name]["clean"] = df_final
                    st.success("Tabla lista para exportar.")

            st.markdown('</div>', unsafe_allow_html=True)

    ready_data = {k: v["clean"] for k, v in st.session_state.master.items() if v.get("clean") is not None}
    if ready_data:
        excel_file = generar_excel_profesional(ready_data)
        st.download_button(
            label="🚀 DESCARGAR EXCEL CON FORMATO CORPORATIVO",
            data=excel_file,
            file_name="REPORTE_FINAL_INVIMA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )